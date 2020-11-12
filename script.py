import openpyxl
import sys
from datetime import datetime
import argparse
import os

SKU_COLUMN_NAME = 'manufacturer_sku'
STOCK_COLUMN_NAME = 'stock'
ETA_COLUMN_NAME = 'eta'
TODAY_DATE = datetime.today().strftime('%d-%m-%Y')
DESTINATION_FOLDER_PATH = '.'
EXCEL_FILETYPE = '.xlsx'

if __name__ == '__main__':
    # get command line arguments
    parser = argparse.ArgumentParser(description='Process an excel file and create a new one in the destination folder')
    parser.add_argument('manufacturer', type=str, help='manufacturer name used for creating new excel file')
    parser.add_argument('inputfile', type=str, help='input excel file')
    args = parser.parse_args()
    
    # get workbook, first sheet is presumed to be the one we want
    workbook = openpyxl.load_workbook(filename=args.inputfile)
    sheet = workbook[workbook.sheetnames[0]]
    print(sheet)

    # rename first column
    original_name = sheet['A1']
    print('Renaming header of first column from ' + (original_name.value if original_name.value else 'N/A') + ' to ' + SKU_COLUMN_NAME)
    sheet['A1'] = SKU_COLUMN_NAME

    # parse first column and strip leading and trailling whitespaces
    print('Parsing first column and stripping whitespaces')
    for row in range(2, sheet.max_row + 1):
        cell_name = 'A{}'.format(row)
        original_value = sheet[cell_name].value
        stripped_value = str(original_value).strip()
        sheet[cell_name].value = stripped_value

    # rename second column
    original_name = sheet['B1']
    print('Renaming header of second column from ' + (original_name.value if original_name.value else 'N/A') + ' to ' + STOCK_COLUMN_NAME)
    sheet['B1'] = STOCK_COLUMN_NAME

    # parse second column, strip leading and trailling whitespaces, check values are numbers
    print('Parsing second column, stripping whitespaces and checking for integer rule violation')
    for row in range(2, sheet.max_row + 1):
        cell_name = 'B{}'.format(row)
        original_value = sheet[cell_name].value
        stripped_value = str(original_value).strip()
        try:
            # try to write the integer
            sheet[cell_name].value = int(stripped_value)
        except ValueError as ve:
            print('Value that can\'t be parsed as an integer found in cell ' + cell_name + '. Error: ' + str(ve))

    # rename third column
    original_name = sheet['C1']
    print('Renaming header of third column from ' + (original_name.value if original_name.value else 'N/A') + ' to ' + ETA_COLUMN_NAME)
    sheet['C1'] = ETA_COLUMN_NAME

    # parse third column and strip leading and trailling whitespaces
    print('Parsing second column, stripping whitespaces and checking for date')
    for row in range(2, sheet.max_row + 1):
        cell_name = 'C{}'.format(row)
        original_value = sheet[cell_name].value
        stripped_value = str(original_value).strip()
        sheet[cell_name].value = stripped_value
        if sheet[cell_name].is_date is False:
            print('Cell ' + cell_name + ' is not a correct date')

    # save new workbook
    if not os.path.exists(DESTINATION_FOLDER_PATH + os.path.sep + TODAY_DATE):
        os.makedirs(DESTINATION_FOLDER_PATH + os.path.sep + TODAY_DATE)
    destination_path = DESTINATION_FOLDER_PATH + os.path.sep + TODAY_DATE
    new_workbook_name = args.manufacturer + '-' + TODAY_DATE
    workbook.save(destination_path + os.path.sep + new_workbook_name + EXCEL_FILETYPE)


